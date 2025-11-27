# _delete_dates_range.py
# 지정한 날짜 범위(start_date ~ end_date)의 열을
# '종목' 시트를 제외한 모든 시트에서 강제로 삭제하는 스크립트

import openpyxl
from datetime import datetime, timedelta
from pathlib import Path


EXCEL_FILE = "_stock_value.xlsx"

# 🔹 삭제할 날짜 범위 직접 지정 (YYYYMMDD)
START_DATE = "20251127"
END_DATE   = "20251127"


def parse_header_date(val):
    """헤더 값을 datetime 으로 변환"""
    if val is None:
        return None

    if isinstance(val, datetime):
        return datetime(val.year, val.month, val.day)

    # Excel serial
    if isinstance(val, (int, float)):
        try:
            base = datetime(1899, 12, 30)
            return base + timedelta(days=int(val))
        except:
            pass

    s = str(val).strip()
    if not s:
        return None

    # 숫자 8자리(YYYYMMDD) 우선 처리
    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) == 8:
        try:
            return datetime.strptime(digits, "%Y%m%d")
        except:
            pass

    # 여러 포맷 시도
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y.%m.%d.", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except:
            pass

    return None


def daterange(start, end):
    """start~end 날짜 리스트 반환"""
    cur = start
    while cur <= end:
        yield cur
        cur += timedelta(days=1)


def main():
    path = Path(EXCEL_FILE)
    if not path.exists():
        print(f"❌ 파일이 없습니다: {EXCEL_FILE}")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)

    # ✅ 삭제 대상 시트: '종목'을 제외한 모든 시트
    target_sheets = [s for s in wb.sheetnames if s != "종목"]

    # 날짜 범위 준비
    start_dt = datetime.strptime(START_DATE, "%Y%m%d")
    end_dt = datetime.strptime(END_DATE, "%Y%m%d")
    delete_dates = set(d.date() for d in daterange(start_dt, end_dt))

    print(f"🗑 삭제할 날짜 범위: {START_DATE} ~ {END_DATE}")
    print(f"   총 {len(delete_dates)}일")
    print(f"   대상 시트: {', '.join(target_sheets)}\n")

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        print(f"\n📄 '{sheet_name}' 시트 처리 중...")

        cols_to_delete = []

        max_col = ws.max_column
        # 1행, 3열부터 날짜 헤더라고 가정
        for col in range(3, max_col + 1):
            raw = ws.cell(row=1, column=col).value
            dt = parse_header_date(raw)
            if dt is None:
                continue

            if dt.date() in delete_dates:
                cols_to_delete.append(col)

        if not cols_to_delete:
            print(f"   → 삭제할 날짜 없음 (패스)")
            continue

        print(f"   → 삭제할 열 번호: {cols_to_delete}")

        # 열 삭제 (뒤에서부터 삭제해야 인덱스가 안 틀림)
        for col in sorted(cols_to_delete, reverse=True):
            ws.delete_cols(col)

        print(f"   ✔ 삭제 완료 ({len(cols_to_delete)}개 열 삭제)")

    wb.save(EXCEL_FILE)
    wb.close()
    print("\n🎉 모든 작업 완료!")


if __name__ == "__main__":
    main()
