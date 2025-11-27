# _delete_latest_day.py
# _stock_value.xlsx에서 "최신 날짜(컬럼)"를 모든 시트에서 한 번에 삭제하는 롤백 스크립트

import sys
from pathlib import Path
import openpyxl


# ------------------------------
# 1. 유틸: 헤더에서 마지막 날짜(YYYYMMDD) 찾기
# ------------------------------
def get_last_date_from_sheet(ws):
    """
    ws(Worksheet)의 1행, 3열 이후 헤더들에서
    숫자 8자리(YYYYMMDD) 형태의 값을 모아
    가장 최신(최대) 날짜 문자열을 반환.
    못 찾으면 None 반환.
    """
    dates = []
    for col in range(3, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        s = "".join(ch for ch in str(v) if ch.isdigit())
        if len(s) == 8:
            dates.append(s)

    if not dates:
        return None

    return max(dates)  # 문자열 비교해도 YYYYMMDD 포맷이면 최신이 가장 큼


# ------------------------------
# 2. 유틸: 특정 날짜(YYYYMMDD)에 해당하는 컬럼 삭제
# ------------------------------
def delete_date_col(ws, ymd):
    """
    ws(Worksheet)에서 1행, 3열 이후 헤더를 돌면서
    숫자 8자리가 ymd와 같은 컬럼을 찾아 delete_cols로 삭제.
    삭제 성공 시 True, 못 찾으면 False 반환.
    """
    target_col = None

    for col in range(3, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        s = "".join(ch for ch in str(v) if ch.isdigit())
        if s == ymd:
            target_col = col
            break

    if target_col is not None:
        ws.delete_cols(target_col)
        return True
    return False


# ------------------------------
# 3. 메인 로직
# ------------------------------
def main():
    excel_path = Path("_stock_value.xlsx")
    if not excel_path.exists():
        print("❌ _stock_value.xlsx 파일을 찾을 수 없습니다.")
        return 1

    wb = openpyxl.load_workbook(excel_path)

    # 🔥 최신일 삭제 대상 시트 목록
    target_sheet_names = [
        "종가",
        "거래량",
        "지수",
        "z20", "z60", "z120",
        "s20", "s60", "s120",
        "gap", "quant",
    ]

    # 실제로 존재하는 시트만 사용
    ws_list = []
    for name in target_sheet_names:
        if name in wb.sheetnames:
            ws_list.append(wb[name])
        else:
            print(f"ℹ️ 워크북에 '{name}' 시트가 없어 건너뜁니다.")

    if not ws_list:
        print("❌ 삭제할 대상 시트가 하나도 없습니다.")
        return 1

    # --------------------------
    # 3-1. 각 시트의 마지막 날짜 수집
    # --------------------------
    last_dates = []  # (시트명, 마지막날짜 or None)
    for ws in ws_list:
        last_ymd = get_last_date_from_sheet(ws)
        last_dates.append((ws.title, last_ymd))

    # 날짜가 하나도 없는 시트 체크
    all_none = all(d is None for _, d in last_dates)
    if all_none:
        print("❌ 어느 시트에서도 날짜 헤더를 찾지 못했습니다.")
        return 1

    # 실제 날짜가 있는 시트만 대상으로 날짜 일관성 체크
    effective = [(name, d) for name, d in last_dates if d is not None]
    unique_dates = {d for _, d in effective}

    if len(unique_dates) != 1:
        print("❌ 시트별 마지막 날짜가 서로 다릅니다. 삭제를 중단합니다.")
        for name, d in last_dates:
            print(f"  - {name}: {d}")
        return 1

    target_date = unique_dates.pop()
    print(f"📅 삭제 대상 날짜(YYYYMMDD): {target_date}")
    print("   (모든 시트의 마지막 날짜가 동일함을 확인했습니다.)")

    # --------------------------
    # 3-2. 각 시트에서 해당 날짜 컬럼 삭제
    # --------------------------
    any_deleted = False
    for ws in ws_list:
        ok = delete_date_col(ws, target_date)
        if ok:
            any_deleted = True
            print(f"  ✅ '{ws.title}' 시트에서 {target_date} 컬럼 삭제 완료")
        else:
            print(f"  ⚠️ '{ws.title}' 시트에서 {target_date} 컬럼을 찾지 못했습니다.")

    if not any_deleted:
        print("❌ 어느 시트에서도 해당 날짜 컬럼을 삭제하지 못했습니다.")
        return 1

    # --------------------------
    # 3-3. 저장
    # --------------------------
    wb.save(excel_path)
    print(f"\n✅ 최신일({target_date}) 삭제 완료: {excel_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
