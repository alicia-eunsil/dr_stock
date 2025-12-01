import openpyxl
from openpyxl.styles import Font, PatternFill
import numpy as np
from decimal import Decimal, ROUND_HALF_UP

# -----------------------------
# 1. 기본 데이터(종가) 읽기
# -----------------------------
def get_close_data(filename):
    dates = []
    stocks = []
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb['종가']
        # 1행: 날짜 (3열부터 끝까지)
        raw_dates = [sheet.cell(row=1, column=col).value for col in range(3, sheet.max_column + 1)]
        for d in raw_dates:
            if d is None:
                continue
            d_str = str(d)
            if len(d_str) == 8 and d_str.isdigit():
                if d_str not in dates:
                    dates.append(d_str)

        # 2행 이후: 종목명, 종목코드, 종가들
        for row in range(2, sheet.max_row + 1):
            name = sheet.cell(row=row, column=1).value
            code = sheet.cell(row=row, column=2).value
            prices = [sheet.cell(row=row, column=col).value for col in range(3, sheet.max_column + 1)]
            prices_numeric = []
            for p in prices:
                try:
                    prices_numeric.append(int(p) if p not in (None, '') else None)
                except (ValueError, TypeError):
                    prices_numeric.append(None)
            stocks.append({'name': name, 'code': code, 'prices': prices_numeric})
    except Exception as e:
        print(f"⚠ 종가 시트 로딩 중 오류: {e}")
    return dates, stocks

# -----------------------------
# 2. std 시트 읽기 (있으면)
# -----------------------------
def get_existing_std_sheet(filename, sheet_name):
    wb = openpyxl.load_workbook(filename)
    if sheet_name not in wb.sheetnames:
        return None, None, None
    sheet = wb[sheet_name]

    # 1행: 날짜 (3열부터)
    raw_dates = [sheet.cell(row=1, column=col).value for col in range(3, sheet.max_column + 1)]
    dates = []
    for d in raw_dates:
        d_str = str(d).strip()
        if len(d_str) == 8 and d_str.isdigit():
            dates.append(d_str)

    # 종목별 STD 값 (필요하면 확장 가능)
    stocks = {}
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row=row, column=1).value
        code = sheet.cell(row=row, column=2).value
        std_values = [sheet.cell(row=row, column=col).value for col in range(3, sheet.max_column + 1)]
        stocks[code] = {'name': name, 'std_values': std_values}

    return dates, stocks, wb

# -----------------------------
# 3. STD 계산 함수
# -----------------------------
def calc_std_value(prices, idx, window_std=20, window_mean=20):
    """
    idx 날짜에서의 STD 값 계산:
    - 먼저 해당 idx에서 20일 롤링 표준편차 σ_t 계산
    - 과거 20일(포함) 각각에 대해 20일 롤링 σ를 계산한 뒤 평균을 내서 평균σ 계산
    - STD = (σ_t / 평균σ) * 100
    데이터 부족 또는 None 포함 시 None 반환
    """
    # STD를 계산하려면 최소 window_std + window_mean - 1 만큼의 데이터 필요
    min_idx = window_std + window_mean - 2  # 예: 20 + 20 - 2 = 38
    if idx < min_idx:
        return None

    # σ_t 및 과거 20일 σ 리스트 계산
    std_list = []
    for j in range(idx - window_mean + 1, idx + 1):  # j: idx-19 ~ idx (20일)
        start = j - window_std + 1
        end = j + 1  # 슬라이스에서 끝 인덱스는 +1
        if start < 0:
            return None
        window_prices = prices[start:end]

        # 가격 중 None 있으면 계산 불가
        if any(p is None for p in window_prices):
            return None

        arr = np.array(window_prices, dtype=float)
        sigma = float(np.std(arr, ddof=0))  # 모표준편차
        std_list.append(sigma)

    if not std_list:
        return None

    std_today = std_list[-1]
    avg_std = sum(std_list) / len(std_list)

    if avg_std == 0:
        return 0

    raw_val = (std_today / avg_std - 1) * 100
    # 소수 둘째 자리까지 반올림
    val = float(Decimal(str(raw_val)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
    return val

# -----------------------------
# 4. std 시트 업데이트(부분 업데이트)
# -----------------------------
def update_std_to_excel(filename, dates, stocks, sheet_name):
    window_std = 20
    window_mean = 20
    min_idx = window_std + window_mean - 2  # STD 계산이 가능한 최소 인덱스 (예: 38)

    existing_dates, existing_std_stocks, wb = get_existing_std_sheet(filename, sheet_name)
    if existing_dates is None:
        return save_all_std_to_excel(filename, dates, stocks, sheet_name, window_std, window_mean)

    sheet = wb[sheet_name]

    # STD가 계산될 수 있는 날짜들만 대상으로
    available_dates = [
        d for i, d in enumerate(dates)
        if i >= min_idx and isinstance(d, str) and len(d) == 8 and d.isdigit()
    ]

    # 새로 추가해야 할 날짜
    new_dates = [d for d in available_dates if d not in existing_dates]

    if not new_dates:
        print(f"✅ {sheet_name.upper()} 업데이트 완료: {filename}")
        print(f"   📊 탭: {sheet_name}")
        print(f"   ➕ 추가된 날짜 수: 0개 (이미 최신 상태)")
        print(f"   📈 종목 수: {len(stocks)}개")
        wb.close()
        return

    # 기존 + 신규 전체 날짜 목록
    all_dates = existing_dates + new_dates

    # 새 날짜의 시작 열 (기존 날짜 수 + 3)
    start_col = len(existing_dates) + 3

    # 1행에 새 날짜 헤더 추가
    for col_offset, date in enumerate(new_dates):
        cell = sheet.cell(row=1, column=start_col + col_offset)
        cell.value = date
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    # 각 종목별로 STD 값 계산하여 채우기
    for row_idx, stock in enumerate(stocks, 2):
        code = stock['code']
        prices = stock['prices']

        for col_offset, new_date in enumerate(new_dates):
            try:
                date_idx = dates.index(new_date)
            except ValueError:
                sheet.cell(row=row_idx, column=start_col + col_offset, value=None)
                continue

            val = calc_std_value(prices, date_idx, window_std=window_std, window_mean=window_mean)
            sheet.cell(row=row_idx, column=start_col + col_offset, value=val)

    # 열 너비 설정
    for col_idx in range(start_col, start_col + len(new_dates)):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12

    wb.save(filename)
    print(f"✅ {sheet_name.upper()} 업데이트 완료: {filename}")
    print(f"   📊 탭: {sheet_name}")
    print(f"   ➕ 추가된 날짜 수: {len(new_dates)}개")
    print(f"   📈 종목 수: {len(stocks)}개")

# -----------------------------
# 5. std 시트 전체 생성(처음 만들 때)
# -----------------------------
def save_all_std_to_excel(filename, dates, stocks, sheet_name, window_std=20, window_mean=20):
    min_idx = window_std + window_mean - 2  # 첫 STD가 나오는 인덱스

    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    # 기존 std 시트가 있으면 삭제 후 새로 생성
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    sheet = wb.create_sheet(sheet_name)

    # 헤더
    sheet.cell(row=1, column=1, value='종목명')
    sheet.cell(row=1, column=2, value='종목코드')

    # 날짜 헤더 (STD가 계산 가능한 날짜들만)
    valid_dates = dates[min_idx:]
    for col, date in enumerate(valid_dates, 3):
        cell = sheet.cell(row=1, column=col)
        cell.value = date
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    sheet.cell(row=1, column=1).font = Font(bold=True)
    sheet.cell(row=1, column=1).fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    sheet.cell(row=1, column=2).font = Font(bold=True)
    sheet.cell(row=1, column=2).fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

    # 데이터 행
    for row_idx, stock in enumerate(stocks, 2):
        sheet.cell(row=row_idx, column=1, value=stock['name'])
        sheet.cell(row=row_idx, column=2, value=stock['code'])
        prices = stock['prices']

        for i in range(min_idx, len(prices)):
            val = calc_std_value(prices, i, window_std=window_std, window_mean=window_mean)
            col_idx = 3 + (i - min_idx)
            sheet.cell(row=row_idx, column=col_idx, value=val)

    # 열 너비
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 12
    for col_idx in range(3, len(valid_dates) + 3):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12

    wb.save(filename)
    print(f"✅ {sheet_name.upper()} 전체 저장 완료: {filename}")
    print(f"   📊 탭: {sheet_name}")
    print(f"   📅 {sheet_name.upper()} 날짜 수: {len(valid_dates)}개")
    print(f"   📈 종목 수: {len(stocks)}개")

# -----------------------------
# 6. main
# -----------------------------
def main():
    filename = '_stock_value.xlsx'  # totalS와 동일 파일 사용
    dates, stocks = get_close_data(filename)

    # std 라는 단일 탭만 생성/업데이트
    sheet_name = 'std'
    update_std_to_excel(filename, dates, stocks, sheet_name)

if __name__ == "__main__":
    main()
