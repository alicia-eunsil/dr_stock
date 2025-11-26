# _stock_dashboard.py
# 추가: 원자료 보기, 더보기 기능 추가
# 추가: 지표별 탭 추가

import streamlit as st
import subprocess
import sys
import time
import pandas as pd
import openpyxl
from pathlib import Path
import os
import bcrypt
from datetime import datetime, date, timedelta

# ======================================
# 0. 인증
# ======================================
ACCESS_CODE_HASH = b"$2b$12$gDBpQYK.g938H.8cNwLeUu/VRidCP1GxqusJiEQzVnvaSrG4CBE6K"

if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

if not st.session_state["authenticated"]:
    st.title("🔒 Access Required")
    st.write("Please enter the access code to open the dashboard.")

    with st.form("auth_form"):
        code = st.text_input("Enter access code", type="password")
        submitted = st.form_submit_button("Submit")

    if submitted:
        if bcrypt.checkpw(code.encode(), ACCESS_CODE_HASH):
            st.session_state["authenticated"] = True
            st.success("Access granted")
            st.rerun()
        else:
            st.error("Invalid code")

    st.stop()

# ======================================
# 페이지 설정
# ======================================
st.set_page_config(page_title="주식 데이터 대시보드", page_icon="📈", layout="wide")

# ======================================
# 상태 변수
# ======================================
if "run_update" not in st.session_state:
    st.session_state.run_update = False
if "data_loaded" not in st.session_state:
    st.session_state.data_loaded = True

# 🔥 종합 탭 날짜 확장용 
if "show_days" not in st.session_state:
    st.session_state.show_days = 10  # 시작: 최근 10일

# 🔥 원자료 탭 날짜 확장용
if "show_days_raw" not in st.session_state:
    st.session_state.show_days_raw = 10  # 시작: 최근 10일

# ======================================
# 날짜 처리 함수
# ======================================
def _to_datetime(v):
    if isinstance(v, (datetime, date)):
        return datetime(v.year, v.month, v.day)

    if isinstance(v, (int, float)):
        base = datetime(1899, 12, 30)
        try:
            return base + timedelta(days=int(v))
        except:
            return None

    s = str(v).strip()
    if not s:
        return None

    for fmt in ("%Y-%m-%d", "%Y.%m.%d.", "%Y.%m.%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except:
            pass

    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) == 8:
        try:
            return datetime.strptime(digits, "%Y%m%d")
        except:
            pass

    return None

# _to_datetime로 바꾼 날짜를 YYYY.MM.DD. 형식 문자열로 변환
def format_excel_date(v):
    dt = _to_datetime(v)
    if dt:
        return dt.strftime("%Y.%m.%d.")
    s = str(v)
    s = s.replace("-", ".").replace("/", ".")
    if not s.endswith("."):
        s += "."
    return s


def _format_z_cell(v):
    val = pd.to_numeric(v, errors="coerce")
    if pd.isna(val):
        return "-"
    out = f"{val:.0f}"
    if val > 100:
        out += " 🔴"
    elif val < -100:
        out += " 🔵"
    return out


def _format_s_cell(v):
    val = pd.to_numeric(v, errors="coerce")
    if pd.isna(val):
        return "-"
    out = f"{val:.0f}"
    if abs(val - 100) < 0.1:
        out += " 🔴"
    elif abs(val - 0) < 0.1:
        out += " 🔵"
    return out

def _format_q_cell(v):
    val = pd.to_numeric(v, errors="coerce")
    if pd.isna(val):
        return "-"
    out = f"{val:.0f}"
    if val > 100:
        out += " 🔴"
    elif val < 25:
        out += " 🔵"
    return out

def _format_price(x):
    """종가(가격)를 세 자리 콤마가 있는 문자열로 변환"""
    try:
        # None, 빈 문자열 처리
        if x is None:
            return ""
        if isinstance(x, str) and x.strip() == "":
            return ""

        v = float(x)
        return f"{v:,.0f}"  # 예: 12345 -> '12,345'
    except:
        # 숫자로 변환 안 되면 빈칸 처리
        return ""

def render_metric_view(indicator_df, selected_labels):
    """
    지표별 탭:
    - 1열: 종목코드
    - 2열: 종목명
    - 3열~: 날짜별 지표값 (S/Z는 이모지 포함, GAP/QUANT는 숫자만)
    """
    st.subheader("📈 지표별 종목 · 일자 조회")

    if indicator_df is None or len(indicator_df) == 0:
        st.warning("⚠️ 지표별 데이터를 불러올 수 없습니다.")
        return

    # -------------------------
    # 0. 선택할 지표 목록 준비
    # -------------------------
    metric_options = ["S20", "S60", "S120",
                      "Z20", "Z60", "Z120",
                      "GAP", "QUANT"]

    # 실제 indicator_df에 존재하는 지표만 남기기
    available = []
    for m in metric_options:
        # indicator_df 컬럼은 (날짜라벨, 지표명) 형태라서,
        # 아무 날짜 하나라도 (lbl, m) 이 존재하면 사용 가능하다고 봄
        if any(((lbl, m) in indicator_df.columns) for lbl in selected_labels):
            available.append(m)

    if not available:
        st.error("indicator_df에 S/Z/GAP/QUANT 관련 컬럼이 없습니다.")
        st.write("현재 indicator_df.columns 예시:", list(indicator_df.columns)[:20])
        return

    metric = st.selectbox("지표를 선택하세요", available, index=0)

    # -------------------------
    # 1. 기본 DF 구성 (종목코드, 종목명 + 날짜별 값)
    # -------------------------
    df_metric = indicator_df[["종목코드", "종목명"]].copy()

    # 선택된 지표에 대해 날짜별 컬럼 추가
    for lbl in selected_labels:
        col_key = (lbl, metric)  # 예: ('2025.01.01.', 'S20')
        if col_key in indicator_df.columns:
            df_metric[lbl] = indicator_df[col_key]
        else:
            df_metric[lbl] = None

    # -------------------------
    # 2. 값 포맷팅 (이모지 포함 / 숫자만)
    # -------------------------
    def _format_plain(v):
        val = pd.to_numeric(v, errors="coerce")
        if pd.isna(val):
            return "-"
        return f"{val:.0f}"

    if metric.startswith("S"):
        formatter = _format_s_cell
    elif metric.startswith("Z"):
        formatter = _format_z_cell
    else:  # GAP, QUANT 등은 기준 없이 숫자만
        formatter = _format_plain

    for lbl in selected_labels:
        if lbl in df_metric.columns:
            df_metric[lbl] = df_metric[lbl].apply(formatter)

    # -------------------------
    # 3. 🔍 필터 옵션 (검색 + 정렬)
    # -------------------------
    st.markdown("### 🔍 필터 옵션 (지표별)")
    c1, c2 = st.columns(2)
    with c1:
        search_metric = st.text_input(
            "🔎 종목명/종목코드 검색",
            key="search_metric"
        )
    with c2:
        sort_metric = st.selectbox(
            "정렬 기준",
            ["종목코드", "종목명"],
            key="sort_metric"
        )

    # 검색 적용
    df_filtered = df_metric.copy()
    if search_metric:
        df_filtered = df_filtered[
            df_filtered["종목명"].astype(str).str.contains(search_metric, case=False)
            | df_filtered["종목코드"].astype(str).str.contains(search_metric, case=False)
        ]

    # 정렬 적용
    df_filtered = df_filtered.sort_values(by=sort_metric).reset_index(drop=True)

    # -------------------------
    # 4. 현재 날짜 범위 표시
    # -------------------------
    if selected_labels:
        oldest_label = selected_labels[0]
        latest_label = selected_labels[-1]
        st.info(
            f"📅 지표별 표시 범위: **{oldest_label} ~ {latest_label}** "
            f"(최근 {len(selected_labels)}일)"
        )

    # -------------------------
    # 5. 테이블 출력
    # -------------------------
    st.markdown(f"### 📋 {metric} · 추이")

    column_config = {
        "종목코드": st.column_config.TextColumn("종목코드", width="small", pinned="left"),
        "종목명": st.column_config.TextColumn("종목명", width="small", pinned="left"),
    }
    for lbl in selected_labels:
        if lbl in df_filtered.columns:
            column_config[lbl] = st.column_config.TextColumn(lbl)

    st.dataframe(
        df_filtered,
        use_container_width=True,
        height=600,
        hide_index=True,
        column_config=column_config,
    )

    # -------------------------
    # 6. ⬅ 과거 10일 더보기(지표별)
    # -------------------------
    # total_days와 show_days는 상단에서 이미 전역으로 관리 중
    global total_days
    if st.button("⬅ 과거 10일 더보기(지표별)", disabled=(total_days <= st.session_state.show_days)):
        st.session_state.show_days = min(st.session_state.show_days + 10, total_days)
        st.rerun()

# ======================================
# 사이드바: 데이터 갱신 버튼
# ======================================
with st.sidebar:
    
    # ✅ _stock_value.xlsx 파일이 있으면 언제든 다운로드 버튼 표시
    excel_path = Path("_stock_value.xlsx")
    if excel_path.exists():
        with open(excel_path, "rb") as f:
            st.download_button(
                label="📥 최신 데이터 다운로드",
                data=f,
                file_name="_stock_value.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_excel",
            )
            
    if st.button("🔄 데이터 갱신 시작"):
        st.session_state.run_update = True
# ======================================
# 데이터 갱신 실행
# ======================================
if st.session_state.run_update:
    with st.sidebar:
        st.subheader("진행 상황")
        pb = st.progress(0)
        msg = st.empty()

    scripts = [
        ("_totalS.py", "S20/S60/S120 계산"),
        ("_totalZ.py", "Z20/Z60/Z120 계산"),
        ("_gap.py", "GAP 계산"),
        ("_quant.py", "QUANT 계산"),
    ]

    for idx, (sc, desc) in enumerate(scripts):
        msg.write(f"{desc} 실행 중...")
        try:
            result = subprocess.run(
                [sys.executable, sc], capture_output=True, text=True, timeout=300
            )
            if result.returncode == 0:
                st.sidebar.success(f"{desc} 완료")
            else:
                st.sidebar.error(f"{desc} 실패")
        except:
            st.sidebar.error(f"{desc} 오류 발생")

        pb.progress((idx + 1) / len(scripts))

    st.session_state.data_loaded = True
    st.session_state.run_update = False
    st.rerun()

# ======================================
# 데이터 로드
# ======================================
excel_files = list(Path(".").glob("_stock_value.xlsx"))
if not excel_files:
    st.error("_stock_value.xlsx 파일을 찾지 못했습니다. "
             "GitHub Actions가 아직 안 돌았거나, "
             "필요하면 왼쪽의 '데이터 갱신 시작' 버튼을 눌러주세요.")
    st.stop()

excel_file = excel_files[0]
wb = openpyxl.load_workbook(excel_file, data_only=True)

# ======================================
# 종목 정보 읽기
# ======================================
stock_info = {}
if "종목" in wb.sheetnames:
    ws = wb["종목"]
    for r in ws.iter_rows(min_row=2, max_col=2):
        name = r[0].value
        code = r[1].value
        if code and name:
            stock_info[code] = name

# ======================================
# 1. 종합(Z20/Z60/.../GAP) 데이터 로딩
# ======================================
sheet_names = ["z20", "z60", "z120", "s20", "s60", "s120", "gap", "quant"]

# 기준 시트 하나 선택 (z20이 됨)
base_ws = None
for s in sheet_names:
    if s in wb.sheetnames:
        base_ws = wb[s]
        break

indicator_df = None
indicator_date_infos = []
total_days = 0

if base_ws:
    max_col = base_ws.max_column

    # 날짜 헤더 수집 (기준: z20 시트 1행, 3열~)
    for col in range(3, max_col + 1):
        raw = base_ws.cell(row=1, column=col).value
        if raw is None:
            continue
        dt = _to_datetime(raw)
        label = format_excel_date(raw)
        indicator_date_infos.append((col, raw, dt, label))

    # 날짜 정렬 (과거 → 최신)
    indicator_date_infos = sorted(
        indicator_date_infos,
        key=lambda x: (x[2] is None, x[2] or datetime.min)
    )

    total_days = len(indicator_date_infos)

    # ➜ 현재 표시할 일수 (최근 N일)
    show_days = min(st.session_state.show_days, total_days)

    # ➜ 가장 최근 show_days개 선택
    start_idx = total_days - show_days
    selected_infos = indicator_date_infos[start_idx:]  # 과거 → 최신
    selected_labels = [lbl for _, _, _, lbl in selected_infos]

    # 날짜 범위 표시용
    oldest_label = selected_infos[0][3]
    latest_label = selected_infos[-1][3]
    indicator_range_msg = (
        f"📅 종합 표시 범위: **{oldest_label} ~ {latest_label}** "
        f"(최근 {show_days}일 / 전체 {total_days}일)"
    )

    # 종목별 데이터 딕셔너리
    data_dict = {code: {"종목코드": code, "종목명": name} for code, name in stock_info.items()}

    # 🔧 시트별로 데이터 가져오기 (열 번호가 아니라 '날짜 문자열'로 매칭!)
    for s in sheet_names:
        if s not in wb.sheetnames:
            continue

        ws = wb[s]
        max_row_s = ws.max_row
        max_col_s = ws.max_column

        # 이 시트의 날짜 → 열번호 매핑 만들기
        label_to_col = {}
        for col in range(3, max_col_s + 1):
            raw = ws.cell(row=1, column=col).value
            if raw is None:
                continue
            lbl = format_excel_date(raw)
            label_to_col[lbl] = col

        # 각 종목별로, 선택된 날짜들에 대해 값 채우기
        for r in range(2, max_row_s + 1):
            code = ws.cell(row=r, column=2).value
            if code not in data_dict:
                continue

            for lbl in selected_labels:
                col_idx = label_to_col.get(lbl)
                if col_idx is None:
                    val = None
                else:
                    val = ws.cell(row=r, column=col_idx).value

                data_dict[code][(lbl, s.upper())] = val

    indicator_df = pd.DataFrame.from_dict(data_dict, orient="index").reset_index(drop=True)

    # ======================================
    # 1-1. 지표별 탭용 df_summary 생성
    #   - 형태: 날짜 / 종목코드 / 종목명 / S20 / S60 / S120 / Z20 / Z60 / Z120 / GAP / QUANT
    # ======================================
    df_summary = None
    if indicator_df is not None:
        metrics_for_summary = ["S20", "S60", "S120",
                               "Z20", "Z60", "Z120",
                               "GAP", "QUANT"]
        records = []

        # indicator_df: 행 = 종목, 열 = ("날짜라벨", "지표명") 튜플
        for _, row in indicator_df.iterrows():
            code = row["종목코드"]
            name = row["종목명"]

            # selected_labels: 현재 화면에 표시 중인 날짜 라벨 리스트
            for lbl in selected_labels:
                rec = {
                    "날짜": lbl,
                    "종목코드": code,
                    "종목명": name,
                }
                for m in metrics_for_summary:
                    col = (lbl, m)
                    if col in indicator_df.columns:
                        rec[m] = row[col]
                    else:
                        rec[m] = None
                records.append(rec)

        df_summary = pd.DataFrame(records)
else:
    indicator_df = None
    df_summary = None

# ======================================
# 2. 원자료(종가) 데이터 로딩 + 확장 기능
# ======================================
close_df = None
close_date_infos = []
total_close_days = 0

if "종가" in wb.sheetnames:
    ws = wb["종가"]
    max_col_c = ws.max_column

    # 날짜 헤더
    close_date_infos = []
    for col in range(3, max_col_c + 1):
        raw = ws.cell(row=1, column=col).value
        if raw is None:
            continue

        # 1) 먼저 _to_datetime으로 시도
        dt = _to_datetime(raw)

        # 2) 그래도 안 되면 숫자 8자리만 뽑아서 날짜로 인식
        if dt is None:
            digits = "".join(ch for ch in str(raw) if ch.isdigit())
            if len(digits) == 8:
                dt = datetime.strptime(digits, "%Y%m%d")

        # 3) 날짜로 못 바꾸면 건너뜀
        if dt is None:
            continue

        # 4) 라벨은 항상 YYYY.MM.DD. 형식으로
        label = dt.strftime("%Y.%m.%d.")
        close_date_infos.append((col, raw, dt, label))

    # 정렬 (과거 → 최신)
    close_date_infos = sorted(
        close_date_infos,
        key=lambda x: (x[2] is None, x[2] or datetime.min)
    )

    total_close_days = len(close_date_infos)

    # 현재 표시할 일수
    show_raw = min(st.session_state.show_days_raw, total_close_days)

    start_idx = total_close_days - show_raw
    selected_close_infos = close_date_infos[start_idx:]  # 과거 → 최신

    oldest_label = selected_close_infos[0][3]
    latest_label = selected_close_infos[-1][3]

    close_range_msg = (
        f"📅 종가 표시 범위: **{oldest_label} ~ {latest_label}** "
        f"(최근 {show_raw}일 / 전체 {total_close_days}일)"
    )

    # 종목별 딕셔너리
    close_dict = {code: {"종목명": name, "종목코드": code} for code, name in stock_info.items()}

    max_row_c = ws.max_row

    for r in range(2, max_row_c + 1):
        code = ws.cell(row=r, column=2).value
        if code not in close_dict:
            continue

        for col_idx, raw, dt, label in selected_close_infos:
            val = ws.cell(row=r, column=col_idx).value
            close_dict[code][label] = val

    close_df = pd.DataFrame.from_dict(close_dict, orient="index").reset_index(drop=True)

    # 🔧 컬럼 이름을 종합 탭과 동일하게 yyyy.mm.dd. 형식으로 통일
    rename_map = {}
    for col in close_df.columns:
        if col in ["종목코드", "종목명"]:
            continue
        rename_map[col] = format_excel_date(col)
    
    close_df = close_df.rename(columns=rename_map)

wb.close()

# ======================================
# 탭 구성
# ======================================
tab_total, tab_metric, tab_raw = st.tabs(["1️⃣ 종합", "2️⃣ 지표별", "3️⃣ 원자료"])

# --------------------------------------
# 종합 탭
# --------------------------------------
with tab_total:
    if indicator_df is None:
        st.warning("⚠️ 종합 데이터를 불러올 수 없습니다.")
    else:
        st.markdown("### 🔍 필터 옵션 (종합)")
        c1, c2 = st.columns(2)
        with c1:
            search = st.text_input("🔎 종목명/종목코드 검색", key="search_total")
        with c2:
            sort_by = st.selectbox("정렬 기준", ["종목코드", "종목명"], key="sort_total")

        # 검색 적용
        df_f = indicator_df.copy()
        if search:
            df_f = df_f[
                df_f["종목명"].astype(str).str.contains(search, case=False) |
                df_f["종목코드"].astype(str).str.contains(search, case=False)
            ]

        df_f = df_f.sort_values(by=sort_by)

        st.info(indicator_range_msg)

        # --------------------------------------
        # 🔥 멀티헤더 생성 (1행: 날짜, 2행: 지표명)
        # --------------------------------------
        metrics = ["Z20", "Z60", "Z120", "S20", "S60", "S120", "GAP", "QUANT"]
        base_cols = ["종목코드", "종목명"]
        df_show = df_f[base_cols].copy()

        col_tuples = [("", "종목코드"), ("", "종목명")]

        # 날짜 × 지표 조합을 모두 생성 (값 없으면 '-'로)
        for lbl in selected_labels:
            for m in metrics:
                key = (lbl, m)
                if key in df_f.columns:
                    df_show[(lbl, m)] = df_f[key]
                else:
                    df_show[(lbl, m)] = "-"
                col_tuples.append((lbl, m))

        df_show.columns = pd.MultiIndex.from_tuples(col_tuples)

        # 🔥 평균 행 추가 (맨 마지막 행)
        avg_row = []
        for col in df_show.columns:
            if col == ("", "종목코드"):
                avg_row.append("AVG")     # 혹은 "" 로 비워도 됨
            elif col == ("", "종목명"):
                avg_row.append("평균")    # 행 라벨
            else:
                lbl, m = col
                key = (lbl, m)
                if key in df_f.columns:
                    # 숫자로 변환 후 평균 계산
                    s = pd.to_numeric(df_f[key], errors="coerce")
                    avg_val = s.mean(skipna=True)
                    avg_row.append(f"{avg_val:.2f}")
                else:
                    avg_row.append(None)

        # 맨 아래에 평균 행 추가
        df_show.loc[len(df_show)] = avg_row

        # Z/S 포맷 적용
        for lbl in selected_labels:
            for m in ["Z20", "Z60", "Z120"]:
                col = (lbl, m)
                if col in df_show.columns:
                    df_show[col] = df_show[col].apply(_format_z_cell)

            for m in ["S20", "S60", "S120"]:
                col = (lbl, m)
                if col in df_show.columns:
                    df_show[col] = df_show[col].apply(_format_s_cell)

            # GAP은 숫자 없으면 '-'로 통일
            col = (lbl, "GAP")
            if col in df_show.columns:
                df_show[col] = df_show[col].apply(
                    lambda v: "-" if pd.isna(pd.to_numeric(v, errors="coerce")) else v
                )
            
            for m in ["QUANT"]:
                col = (lbl, m)
                if col in df_show.columns:
                    df_show[col] = df_show[col].apply(_format_q_cell)

        df_show = df_show.set_index([("", "종목코드"), ("", "종목명")])

        st.dataframe(
            df_show,
            use_container_width=True,
            height=600,
        )

        # 🔥 과거 확장 버튼
        if st.button("⬅ 과거 10일 더보기(종합)", disabled=(total_days <= st.session_state.show_days)):
            st.session_state.show_days = min(st.session_state.show_days + 10, total_days)
            st.rerun()

# --------------------------------------
# 원자료 탭
# --------------------------------------
with tab_raw:
    if close_df is None:
        st.warning("⚠️ 원자료(종가) 데이터를 불러올 수 없습니다.")
    else:
        st.markdown("### 🔍 필터 옵션 (원자료)")
        r1, r2 = st.columns(2)
        with r1:
            search_raw = st.text_input("🔎 종목명/종목코드 검색", key="search_raw")
        with r2:
            sort_raw = st.selectbox("정렬 기준", ["종목코드", "종목명"], key="sort_raw")

        df_raw = close_df.copy()

        if search_raw:
            df_raw = df_raw[
                df_raw["종목코드"].astype(str).str.contains(search_raw, case=False) |
                df_raw["종목명"].astype(str).str.contains(search_raw, case=False)
            ]

        df_raw = df_raw.sort_values(by=sort_raw)

        st.info(close_range_msg)

        # 표시 조건 설정
        date_cols = [c for c in df_raw.columns if c not in ["종목코드", "종목명"]]

        # 🔒 컬럼 순서 고정: 종목코드 → 종목명 → 날짜들
        df_raw = df_raw[["종목코드", "종목명"] + date_cols]

        # 🔥 세 자리 콤마 포맷 적용 (모든 날짜 컬럼에)
        for c in date_cols:
            df_raw[c] = df_raw[c].apply(_format_price)

        # 컬럼 설정: 종목코드/종목명은 왼쪽 고정, 날짜들은 텍스트 컬럼
        column_config = {
            "종목코드": st.column_config.TextColumn("종목코드", width="small", pinned="left"),
            "종목명": st.column_config.TextColumn("종목명", width="small", pinned="left"),
        }

        # 날짜 컬럼은 문자열(콤마 포함)이라 TextColumn으로 표시
        for c in date_cols:
            column_config[c] = st.column_config.TextColumn(c)

        st.dataframe(
            df_raw,
            use_container_width=True,
            height=600,
            hide_index=True,
            column_config=column_config,
        )

        # 🔥 과거 확장 버튼
        if st.button("⬅ 과거 10일 더보기(종가)", disabled=(total_close_days <= st.session_state.show_days_raw)):
            st.session_state.show_days_raw = min(st.session_state.show_days_raw + 10, total_close_days)
            st.rerun()

with tab_metric:
    if indicator_df is None:
        st.warning("⚠️ 지표별 데이터를 불러올 수 없습니다.")
    else:
        render_metric_view(indicator_df, selected_labels)

st.markdown("---")
st.caption("Created by Alicia")
