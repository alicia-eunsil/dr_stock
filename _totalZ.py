import openpyxl
from openpyxl.styles import Font, PatternFill
import numpy as np
from decimal import Decimal, ROUND_HALF_UP

# Z 점수 계산 함수 (윈도우 길이만 다름)
def calc_z(prices, window):
    arr = [p for p in prices if p is not None]
    if len(arr) < window:
        return None
    arr = arr[-window:]
    mean = np.mean(arr)
    std = np.std(arr, ddof=1)
    if std == 0:
        return 0
    z = (arr[-1] - mean) / std
    val = 50 * z
    score = int(Decimal(str(val)).to_integral_value(rounding=ROUND_HALF_UP))
    return score

def unique_preserve_order(seq):
    seen = set()
    result = []
    for x in seq:
        x_str = str(x).strip()
        if x_str not in seen:
            seen.add(x_str)
            result.append(x_str)
    return result

def get_close_data(filename):
    dates = []
    stocks = []
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb['종가']
        raw_dates = [sheet.cell(row=1, column=col).value for col in range(3, sheet.max_column+1)]
        for d in raw_dates:
            if d is None:
                continue
            d_str = str(d)
            if len(d_str) == 8 and d_str.isdigit():
                if d_str not in dates:
                    dates.append(d_str)
        for row in range(2, sheet.max_row+1):
            name = sheet.cell(row=row, column=1).value
            code = sheet.cell(row=row, column=2).value
            prices = [sheet.cell(row=row, column=col).value for col in range(3, sheet.max_column+1)]
            prices_numeric = []
            for p in prices:
                try:
                    prices_numeric.append(int(p) if p not in (None, '') else None)
                except (ValueError, TypeError):
                    prices_numeric.append(None)
            stocks.append({'name': name, 'code': code, 'prices': prices_numeric})
    except Exception as e:
        pass
    return dates, stocks

def get_existing_z_sheet(filename, sheet_name):
    wb = openpyxl.load_workbook(filename)
    if sheet_name not in wb.sheetnames:
        return None, None, None
    sheet = wb[sheet_name]
    raw_dates = [sheet.cell(row=1, column=col).value for col in range(3, sheet.max_column+1)]
    dates = []
    for d in raw_dates:
        d_str = str(d).strip()
        if len(d_str) == 8 and d_str.isdigit():
            dates.append(d_str)
    stocks = {}
    for row in range(2, sheet.max_row+1):
        name = sheet.cell(row=row, column=1).value
        code = sheet.cell(row=row, column=2).value
        z_values = [sheet.cell(row=row, column=col).value for col in range(3, sheet.max_column+1)]
        stocks[code] = {'name': name, 'z_values': z_values}
    return dates, stocks, wb

def update_z_to_excel(filename, dates, stocks, window, sheet_name):
    existing_dates, existing_z_stocks, wb = get_existing_z_sheet(filename, sheet_name)
    if existing_dates is None:
        return save_all_z_to_excel(filename, dates, stocks, window, sheet_name)
    
    sheet = wb[sheet_name]
    available_dates = [d for d in dates[window-1:] if isinstance(d, str) and len(d) == 8 and d.isdigit()]
    new_dates = [d for d in available_dates if d not in existing_dates]
    
    if not new_dates:
        print(f"✅ {sheet_name.upper()} 업데이트 완료: {filename}")
        print(f"   📊 탭: {sheet_name}")
        print(f"   ➕ 추가된 날짜 수: 0개 (이미 최신 상태)")
        print(f"   📈 종목 수: {len(stocks)}개")
        wb.close()
        return
    all_dates = existing_dates + new_dates
    start_col = len(existing_dates) + 3
    for col_offset, date in enumerate(new_dates):
        cell = sheet.cell(row=1, column=start_col + col_offset)
        cell.value = date
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    for row_idx, stock in enumerate(stocks, 2):
        code = stock['code']
        prices = stock['prices']
        for col_offset, new_date in enumerate(new_dates):
            try:
                date_idx = dates.index(new_date)
            except ValueError:
                sheet.cell(row=row_idx, column=start_col + col_offset, value=None)
                continue
            if date_idx >= window - 1:
                window_prices = prices[date_idx - window + 1:date_idx + 1]
                if None in window_prices:
                    val = None
                else:
                    val = calc_z(window_prices, window)
            else:
                val = None
            sheet.cell(row=row_idx, column=start_col + col_offset, value=val)
    for col_idx in range(start_col, start_col + len(new_dates)):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12
    wb.save(filename)
    print(f"✅ {sheet_name.upper()} 업데이트 완료: {filename}")
    print(f"   📊 탭: {sheet_name}")
    print(f"   ➕ 추가된 날짜 수: {len(new_dates)}개")
    print(f"   📈 종목 수: {len(stocks)}개")

def save_all_z_to_excel(filename, dates, stocks, window, sheet_name):
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    sheet = wb.create_sheet(sheet_name)
    sheet.cell(row=1, column=1, value='종목명')
    sheet.cell(row=1, column=2, value='종목코드')
    for col, date in enumerate(dates[window-1:], 3):
        cell = sheet.cell(row=1, column=col)
        cell.value = date
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    sheet.cell(row=1, column=1).font = Font(bold=True)
    sheet.cell(row=1, column=1).fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    sheet.cell(row=1, column=2).font = Font(bold=True)
    sheet.cell(row=1, column=2).fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    for row_idx, stock in enumerate(stocks, 2):
        sheet.cell(row=row_idx, column=1, value=stock['name'])
        sheet.cell(row=row_idx, column=2, value=stock['code'])
        prices = stock['prices']
        for i in range(window-1, len(prices)):
            window_prices = prices[i-window+1:i+1]
            if None in window_prices:
                val = None
            else:
                val = calc_z(window_prices, window)
            sheet.cell(row=row_idx, column=3+i-(window-1), value=val)
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 12
    for col_idx in range(3, len(dates[window-1:]) + 3):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12
    wb.save(filename)
    print(f"✅ {sheet_name.upper()} 전체 저장 완료: {filename}")
    print(f"   📊 탭: {sheet_name}")
    print(f"   📅 {sheet_name.upper()} 날짜 수: {len(dates[window-1:])}개")
    print(f"   📈 종목 수: {len(stocks)}개")

def main():
    filename = '_stock_value.xlsx'
    dates, stocks = get_close_data(filename)
    for window, sheet_name in zip([20, 60, 120], ['z20', 'z60', 'z120']):
        update_z_to_excel(filename, dates, stocks, window, sheet_name)

if __name__ == "__main__":
    main()
