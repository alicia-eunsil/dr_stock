'''엑셀에 종가, 거래량 탭이 없으면 최근 100일치 값 가져옴. 있으면 마지막날 확인해서 최신까지의 값 가져와서 저장함'''
import json
import requests
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill
import time

def load_api_secrets(file_path='secrets.json'):
    """API 키와 시크릿을 파일에서 로드"""
    try:
        with open(file_path, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"에러: {file_path} 파일을 찾을 수 없습니다.")
        return None

def get_token(api_key, api_secret, domain):
    """한국투자증권 API 토큰 발급 요청"""
    url = f"{domain}/oauth2/tokenP"
    
    headers = {
        "content-type": "application/json",
        "appKey": api_key,
        "appSecret": api_secret
    }
    
    data = {
        "grant_type": "client_credentials",
        "appkey": api_key,
        "appsecret": api_secret
    }
    
    try:
        resp = requests.post(url, headers=headers, json=data)
        
        if resp.status_code != 200:
            print(f"❌ 토큰 요청 실패: HTTP {resp.status_code}")
            return None
            
        token_data = resp.json()
        if not token_data or 'access_token' not in token_data:
            print("❌ 토큰 정보가 응답에 없습니다")
            return None
            
        print("✅ 토큰 발급 성공!")
        return token_data
        
    except requests.exceptions.RequestException as e:
        print(f"❌ 토큰 요청 실패: {str(e)}")
        if hasattr(e, 'response') and e.response is not None:
            print(f"서버 응답: {e.response.text}")
        return None

def fetch_stock_daily_history(access_token, domain, symbol, start_date, end_date, app_key=None, app_secret=None):
    """한국투자증권의 일별 시세 조회 API를 사용하여 종목의 OHLC 데이터를 가져옵니다."""
    endpoint = f"{domain}/uapi/domestic-stock/v1/quotations/inquire-daily-itemchartprice"
    
    params = {
        "FID_COND_MRKT_DIV_CODE": "J",   # 주식 시장 구분
        "FID_INPUT_ISCD": symbol,         # 종목코드
        "FID_PERIOD_DIV_CODE": "D",       # 기간 구분 (일/주/월)
        "FID_ORG_ADJ_PRC": "1",          # 수정주가 여부
        "FID_INPUT_DATE_1": start_date,   # 조회 시작일
        "FID_INPUT_DATE_2": end_date,     # 조회 종료일
        "FID_COMP_ICD": symbol,          # 종목코드
    }
    
    headers = {
        "content-type": "application/json; charset=utf-8",
        "authorization": f"Bearer {access_token}",
        "appkey": app_key,
        "appsecret": app_secret,
        "tr_id": "FHKST03010100",     # 주식 일별 시세
        "custtype": "P",              # 고객타입: 개인
        "seq_no": "0",               # 시퀀스 번호
        "locale": "ko_KR",          # 언어 설정
    }
    
    try:
        resp = requests.get(endpoint, headers=headers, params=params, timeout=10)
        
        if resp.status_code != 200:
            print(f"❌ HTTP {resp.status_code} 에러: {resp.text}")
            return None
        
        data = resp.json()
        if not data or 'output2' not in data or not data['output2']:
            print("❌ 데이터가 비어있습니다")
            return None
            
        # 일별 데이터 리스트 변환
        daily_data = []
        for item in data['output2']:
            daily_data.append({
                'date': item.get('stck_bsop_date', ''),
                'open': int(item.get('stck_oprc', '0')),
                'high': int(item.get('stck_hgpr', '0')),
                'low': int(item.get('stck_lwpr', '0')),
                'close': int(item.get('stck_clpr', '0')),
                'volume': int(item.get('acml_vol', '0'))
            })
            
        return daily_data
        
    except Exception as e:
        print(f"❌ 데이터 조회 중 에러: {str(e)}")
        return None

def load_stock_list(filename="*stock_value.xlsx"):
    """Excel 파일에서 종목 목록을 읽어옵니다."""
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        
        stocks = []
        for row in sheet.iter_rows(min_row=2):  # 헤더 제외
            if row[0].value and row[1].value:  # 종목명과 코드가 모두 있는 경우만
                # 종목코드를 문자열로 변환하고 6자리로 맞춤
                code = str(row[1].value).strip()
                code = code.zfill(6)  # 6자리 문자열로 변환 (앞에 0 채움)
                
                stocks.append({
                    'name': row[0].value,
                    'code': code
                })
                
        print("\n읽어온 종목 목록:")
        for stock in stocks:
            print(f"  • {stock['name']} (코드: {stock['code']})")
            
        return stocks
        
    except Exception as e:
        print(f"\n❌ Excel 파일 읽기 실패: {str(e)}")
        return None

def save_history_to_excel(data_list, filename="*stock_value.xlsx"):
    """
    각 종목의 일별 OHLC 데이터를 시가/고가/저가/종가/거래량 탭으로 나누어 저장합니다.
    각 탭의 행=종목, 열=일자 매트릭스 형식으로 저장됩니다.
    """
    # 기존 파일이 있으면 로드, 없으면 새로 생성
    try:
        wb = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    # 모든 종목의 날짜를 수집하여 정렬 (과거 → 최신순)
    all_dates = set()
    for stock_data in data_list:
        if stock_data['history']:
            for daily in stock_data['history']:
                all_dates.add(daily['date'])

    # 날짜 정렬 (과거 → 최신순)
    sorted_dates = sorted(list(all_dates))

    if not sorted_dates:
        print("\n❌ 저장할 데이터가 없습니다.")
        return

    # 5개 시트 생성: 시가, 고가, 저가, 종가, 거래량
    sheet_configs = [
        ('시가', 'open'),
        ('고가', 'high'),
        ('저가', 'low'),
        ('종가', 'close'),
        ('거래량', 'volume')
    ]

    for sheet_name, field_name in sheet_configs:
        # 기존 시트가 있으면 기존 데이터 읽기, 없으면 새로 생성
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # 날짜 헤더를 int로 통일
            existing_dates = []
            for col in range(3, sheet.max_column+1):
                val = sheet.cell(row=1, column=col).value
                try:
                    existing_dates.append(int(val))
                except Exception:
                    continue
            existing_data = {}
            for row in range(2, sheet.max_row+1):
                name = sheet.cell(row=row, column=1).value
                code = sheet.cell(row=row, column=2).value
                if not name or not code:
                    continue
                code = str(code).zfill(6)
                values = {}
                for col, date in enumerate(existing_dates, 3):
                    values[str(date)] = sheet.cell(row=row, column=col).value
                existing_data[code] = {'name': name, 'values': values}
        else:
            sheet = wb.create_sheet(sheet_name)
            existing_dates = []
            existing_data = {}

        # 새로 가져온 날짜와 기존 날짜 합치기
        merged_dates = set(existing_dates)
        for stock_data in data_list:
            if stock_data['history']:
                for daily in stock_data['history']:
                    try:
                        merged_dates.add(int(daily['date']))
                    except Exception:
                        continue
        sorted_dates = sorted(list(merged_dates))
        new_dates = set(sorted_dates) - set(existing_dates)

        # 헤더 행 작성
        sheet.cell(row=1, column=1, value='종목명')
        sheet.cell(row=1, column=2, value='종목코드')
        for col, date in enumerate(sorted_dates, 3):
            cell = sheet.cell(row=1, column=col)
            cell.value = date  # int로 저장
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        sheet.cell(row=1, column=1).font = Font(bold=True)
        sheet.cell(row=1, column=1).fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        sheet.cell(row=1, column=2).font = Font(bold=True)
        sheet.cell(row=1, column=2).fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

        # 각 종목별로 행 추가 (기존+신규)
        all_codes = set(existing_data.keys())
        for stock_data in data_list:
            all_codes.add(stock_data['code'])
        for row_idx, code in enumerate(sorted(all_codes), 2):
            # 종목명
            if code in existing_data:
                name = existing_data[code]['name']
            else:
                name = next((s['name'] for s in data_list if s['code'] == code), code)
            sheet.cell(row=row_idx, column=1, value=name)
            sheet.cell(row=row_idx, column=2, value=code)

            # 기존 값 복원
            values = existing_data.get(code, {}).get('values', {})
            # 신규 값 병합
            new_values = {}
            stock_hist = next((s for s in data_list if s['code'] == code), None)
            if stock_hist and stock_hist['history']:
                for daily in stock_hist['history']:
                    try:
                        new_values[str(int(daily['date']))] = daily[field_name]
                    except Exception:
                        continue

            # 날짜별로 값 입력 (신규값 우선, 없으면 기존값)
            for col, date in enumerate(sorted_dates, 3):
                value = new_values.get(str(date), values.get(str(date), ''))
                sheet.cell(row=row_idx, column=col, value=value)

        # 열 너비 자동 조정
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 12
        for col_idx in range(3, len(sorted_dates) + 3):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 12

    wb.save(filename)
    print(f"\n✅ 엑셀 파일 저장 완료: {filename}")
    # 아래 상세 출력은 비활성화
    # print(f"   📊 생성된 탭: 시가, 고가, 저가, 종가, 거래량")
    # print(f"   📅 전체 일자 수: {len(sorted_dates)}일")
    # print(f"   ➕ 이번에 추가된 날짜 수: {len(new_dates)}일")
    # print(f"   📈 종목 수: {len(data_list)}개")

def get_latest_date_from_sheet(filename, sheet_name):
    try:
        wb = openpyxl.load_workbook(filename)
        if sheet_name not in wb.sheetnames:
            return None
        sheet = wb[sheet_name]
        # 첫 행은 헤더, 3번째 열부터 날짜
        dates = [sheet.cell(row=1, column=col).value for col in range(3, sheet.max_column+1)]
        # 날짜가 문자열이면 정렬을 위해 datetime으로 변환
        dates_dt = []
        for d in dates:
            try:
                dates_dt.append(datetime.strptime(str(d), '%Y%m%d'))
            except Exception:
                pass
        if not dates_dt:
            return None
        latest = max(dates_dt)
        return latest.strftime('%Y%m%d')
    except Exception as e:
        print(f"❌ 날짜 추출 에러: {e}")
        return None

def main():
    print(f"\n=== 한국투자증권 API 주식 시세 히스토리 조회 ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')}) ===")
    
    # API 설정 로드
    secrets = load_api_secrets()
    if not secrets:
        return
    
    app_key = secrets.get('api_key')  # api_key로 변경
    app_secret = secrets.get('api_secret')  # api_secret로 변경
    domain = secrets.get('domain', 'https://openapi.koreainvestment.com:9443')
    
    # 종목 목록 로드
    stocks = load_stock_list()
    if not stocks:
        return
    
    # 토큰 발급
    print("\n🔄 토큰 발급 요청 중...")
    token_data = get_token(app_key, app_secret, domain)
    if not token_data:
        print("\n❌ 토큰 발급 실패")
        return
    
    access_token = token_data['access_token']
    filename = "*stock_value.xlsx"
    # 종가/거래량 탭에서 최신 날짜 확인
    latest_close = get_latest_date_from_sheet(filename, "종가")
    latest_amount = get_latest_date_from_sheet(filename, "거래량")
    # 시작일 결정
    if latest_close and latest_amount:
        start_dt = max(latest_close, latest_amount)
        start_dt = datetime.strptime(start_dt, '%Y%m%d') + timedelta(days=1)
        start_date = start_dt.strftime('%Y%m%d')
        print(f"\n📅 추가 조회: {start_date} ~ {datetime.now().strftime('%Y%m%d')}")
    else:
        end_date = datetime.now()
        start_date = (end_date - timedelta(days=100)).strftime('%Y%m%d')
        end_date = end_date.strftime('%Y%m%d')
        print(f"\n📅 전체 조회: {start_date} ~ {end_date}")
    end_date = datetime.now().strftime('%Y%m%d')
    print(f"\n총 {len(stocks)}개 종목에 대해 조회합니다...")
    data_list = []
    for i, stock in enumerate(stocks, 1):
        # ...기존 코드...
        print(f"  [{i}/{len(stocks)}] {stock['name']}({stock['code']}) ...", end='')
        history = fetch_stock_daily_history(
            access_token, 
            domain,
            stock['code'],
            start_date,
            end_date,
            app_key,
            app_secret
        )
        # 이미 저장된 마지막 날짜 이후 데이터만 필터링
        latest = latest_close if stock['code'] == stocks[0]['code'] else None
        if stock['code'] == '010140':
            latest = latest_close
        if latest:
            filtered_history = [d for d in history if int(d['date']) > int(latest)]
        else:
            filtered_history = history
        # ...기존 코드...
        if filtered_history:
            print(f"성공 ({len(filtered_history)}일)")
            data_list.append({
                'name': stock['name'],
                'code': stock['code'],
                'history': filtered_history
            })
        else:
            print("실패")
        time.sleep(1)
    if data_list:
        save_history_to_excel(data_list)
    else:
        print("\n❌ 저장할 데이터가 없습니다.")

if __name__ == "__main__":
    main()
